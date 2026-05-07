from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
import re

from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from app.extensions import db
from app.models import (
    Classroom,
    CourseEnrollment,
    CourseSchedule,
    Group,
    Student,
    TeacherGroupLink,
    User,
)


PASSWORD = "password"
SOURCE_NAME = "real_schedule_xlsx"

TEACHERS = {
    "volchenko_elizaveta": {
        "name": "Вольченко Елизавета Николаевна",
        "email": "volchenko.elizaveta@educam.local",
        "room": "4",
        "aliases": ("вольченко елизавета",),
    },
    "kitaev_viktor": {
        "name": "Китаев Виктор Алексеевич",
        "email": "kitaev.viktor@educam.local",
        "room": "2",
        "aliases": ("китаев виктор",),
    },
    "mikhailova_alena": {
        "name": "Михайлова Алёна Сергеевна",
        "email": "mikhailova.alena@educam.local",
        "room": "1",
        "aliases": ("михайлова алёна", "михайлова алена"),
    },
    "kireeva_irina": {
        "name": "Киреева Ирина Олеговна",
        "email": "kireeva.irina@educam.local",
        "room": "5",
        "aliases": ("киреева ирина",),
    },
    "volchenko_margarita": {
        "name": "Вольченко Маргарита Валентиновна",
        "email": "volchenko.margarita@educam.local",
        "room": "6",
        "aliases": ("вольченко маргарита",),
    },
    "maksimova_anna": {
        "name": "Максимова Анна Викторовна",
        "email": "maksimova.anna@educam.local",
        "room": "9",
        "aliases": ("максимова анна",),
    },
    "kravchuk_ekaterina": {
        "name": "Кравчук Екатерина Владимировна",
        "email": "kravchuk.ekaterina@educam.local",
        "room": "11",
        "aliases": ("кравчук екатерина",),
    },
}

IT_WORDS = (
    "it",
    "айти",
    "информ",
    "программ",
    "python",
    "питон",
    "робот",
    "роблокс",
    "roblox",
    "сайт",
    "web",
    "3d",
    "3д",
    "дрон",
)

DAYS = (
    "понедельник",
    "вторник",
    "среда",
    "четверг",
    "пятница",
    "суббота",
    "воскресенье",
)


@dataclass
class ImportSummary:
    teachers: Counter
    rooms: Counter
    schedules: int
    enrollments: int
    students: int
    student_accounts: dict[str, str]
    teacher_accounts: dict[str, str]


def import_real_schedule(workbook_path: str | Path) -> ImportSummary:
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    ws = wb["Май 26"] if "Май 26" in wb.sheetnames else wb.worksheets[2]

    db.create_all()
    classrooms = ensure_classrooms()
    teacher_users = ensure_teachers()

    seen_students: dict[str, User] = {}
    teacher_counts: Counter = Counter()
    room_counts: Counter = Counter()
    schedule_ids = set()
    enrollment_ids = set()

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        teacher_key = match_teacher(row_value(row, 8))
        if not teacher_key:
            continue

        student_name = clean_text(row_value(row, 1))
        if not student_name:
            continue

        subject = clean_text(row_value(row, 3)) or clean_text(row_value(row, 9)) or "Занятие"
        direction = clean_text(row_value(row, 9))
        level = clean_text(row_value(row, 4))
        status = normalize_status(row_value(row, 5))

        teacher_info = TEACHERS[teacher_key]
        room_number = choose_room(teacher_key, subject, direction)
        classroom = classrooms[room_number]
        teacher = teacher_users[teacher_key]
        student = seen_students.get(normal_key(student_name)) or ensure_student(student_name)
        seen_students[normal_key(student_name)] = student

        for day_col, start_col, end_col in ((10, 11, 12), (13, 14, 15), (16, 17, 18), (19, 20, 21), (22, 23, 24), (25, 26, 27)):
            day = normalize_day(row_value(row, day_col))
            starts_at = normalize_time(row_value(row, start_col))
            ends_at = normalize_time(row_value(row, end_col))
            if not day or not starts_at or not ends_at:
                continue

            group = ensure_group(teacher_info["name"], subject, direction, level, day, starts_at)
            ensure_teacher_group(teacher.id, group.id)
            schedule = ensure_schedule(
                teacher_id=teacher.id,
                group_id=group.id,
                classroom_id=classroom.id,
                title=build_title(subject, direction, level),
                subject=subject,
                direction=direction,
                day=day,
                starts_at=starts_at,
                ends_at=ends_at,
                row_index=row_index,
            )
            enrollment = ensure_enrollment(schedule.id, student.id, status, row_index)
            teacher_counts[teacher_info["name"]] += 1
            room_counts[classroom.name] += 1
            schedule_ids.add(schedule.id)
            enrollment_ids.add(enrollment.id)

            ensure_student_profile(student, group.id)

    db.session.commit()
    return ImportSummary(
        teachers=teacher_counts,
        rooms=room_counts,
        schedules=len(schedule_ids),
        enrollments=len(enrollment_ids),
        students=len(seen_students),
        student_accounts={user.name: user.email for user in sorted(seen_students.values(), key=lambda item: item.name)},
        teacher_accounts={info["name"]: info["email"] for info in TEACHERS.values()},
    )


def ensure_classrooms() -> dict[str, Classroom]:
    rooms = {
        "1": "Кабинет 1",
        "2": "Кабинет 2 · IT",
        "4": "Кабинет 4",
        "5": "Кабинет 5",
        "6": "Кабинет 6",
        "9": "Кабинет 9",
        "11": "Кабинет 11",
    }
    result = {}
    for number, name in rooms.items():
        classroom = Classroom.query.filter_by(number=number).first()
        if not classroom:
            classroom = Classroom(number=number, name=name)
            db.session.add(classroom)
        else:
            classroom.name = name
        result[number] = classroom
    db.session.flush()
    return result


def ensure_teachers() -> dict[str, User]:
    result = {}
    for key, info in TEACHERS.items():
        user = User.query.filter_by(email=info["email"]).first()
        if not user:
            user = User(email=info["email"], name=info["name"], role="teacher", password_hash=generate_password_hash(PASSWORD))
            db.session.add(user)
        else:
            user.name = info["name"]
            user.role = "teacher"
        result[key] = user
    db.session.flush()
    return result


def ensure_student(name: str) -> User:
    email = f"{slugify_name(name)}@educam.local"
    user = User.query.filter_by(email=email).first()
    if not user:
        user = User(email=email, name=name, role="student", password_hash=generate_password_hash(PASSWORD))
        db.session.add(user)
    else:
        user.name = name
        user.role = "student"
    db.session.flush()
    return user


def ensure_student_profile(user: User, group_id: int) -> None:
    profile = Student.query.filter_by(user_id=user.id).first()
    if not profile:
        db.session.add(Student(user_id=user.id, group_id=group_id, full_name=user.name))
    else:
        profile.full_name = user.name
        if not profile.group_id:
            profile.group_id = group_id


def ensure_group(teacher_name: str, subject: str, direction: str, level: str, day: str, starts_at: str) -> Group:
    group_name = " · ".join(part for part in (subject, direction, level, day, starts_at) if part)
    group_name = f"{teacher_name.split()[0]} · {group_name}"[:120]
    group = Group.query.filter_by(name=group_name).first()
    if not group:
        group = Group(name=group_name, description=f"{teacher_name}. Реальное расписание из Excel.")
        db.session.add(group)
        db.session.flush()
    return group


def ensure_teacher_group(teacher_id: int, group_id: int) -> None:
    if not TeacherGroupLink.query.filter_by(teacher_id=teacher_id, group_id=group_id).first():
        db.session.add(TeacherGroupLink(teacher_id=teacher_id, group_id=group_id))


def ensure_schedule(
    teacher_id: int,
    group_id: int,
    classroom_id: int,
    title: str,
    subject: str,
    direction: str,
    day: str,
    starts_at: str,
    ends_at: str,
    row_index: int,
) -> CourseSchedule:
    schedule = CourseSchedule.query.filter_by(
        teacher_id=teacher_id,
        group_id=group_id,
        classroom_id=classroom_id,
        day_of_week=day,
        starts_at=starts_at,
        ends_at=ends_at,
    ).first()
    if not schedule:
        schedule = CourseSchedule(
            teacher_id=teacher_id,
            group_id=group_id,
            classroom_id=classroom_id,
            title=title,
            subject=subject,
            direction=direction,
            day_of_week=day,
            starts_at=starts_at,
            ends_at=ends_at,
            source=SOURCE_NAME,
            source_row=row_index,
        )
        db.session.add(schedule)
        db.session.flush()
    else:
        schedule.title = title
        schedule.subject = subject
        schedule.direction = direction
        schedule.source_row = row_index
    return schedule


def ensure_enrollment(schedule_id: int, student_id: int, status: str, row_index: int) -> CourseEnrollment:
    enrollment = CourseEnrollment.query.filter_by(schedule_id=schedule_id, student_id=student_id).first()
    if not enrollment:
        enrollment = CourseEnrollment(schedule_id=schedule_id, student_id=student_id, status=status, source_row=row_index)
        db.session.add(enrollment)
        db.session.flush()
    else:
        enrollment.status = status
        enrollment.source_row = row_index
    return enrollment


def match_teacher(value) -> str | None:
    teacher = normal_key(value)
    if not teacher:
        return None
    for key, info in TEACHERS.items():
        if any(alias in teacher for alias in info["aliases"]):
            return key
    return None


def choose_room(teacher_key: str, subject: str, direction: str) -> str:
    if teacher_key == "volchenko_elizaveta" and is_it_course(subject, direction):
        return "2"
    return TEACHERS[teacher_key]["room"]


def is_it_course(*parts: str) -> bool:
    text = " ".join(parts).lower()
    return any(word in text for word in IT_WORDS)


def normalize_day(value) -> str:
    text = normal_key(value)
    if not text:
        return ""
    for day in DAYS:
        if day in text:
            return day.capitalize()
    return clean_text(value)


def normalize_time(value) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if not text or text in {"0", "-", "—"}:
        return ""
    match = re.search(r"(\d{1,2})[:.](\d{2})", text)
    if match:
        return f"{int(match.group(1)):02d}:{match.group(2)}"
    return text


def normalize_status(value) -> str:
    text = normal_key(value)
    if "уш" in text or "раст" in text:
        return "inactive"
    if "замор" in text:
        return "paused"
    return "active"


def build_title(subject: str, direction: str, level: str) -> str:
    return " · ".join(part for part in (subject, direction, level) if part) or "Занятие"


def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return re.sub(r"\s+", " ", text)


def normal_key(value) -> str:
    return clean_text(value).lower().replace("ё", "е")


def row_value(row, index: int):
    return row[index] if index < len(row) else None


def slugify_name(name: str) -> str:
    translit = {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "е": "e",
        "ё": "e",
        "ж": "zh",
        "з": "z",
        "и": "i",
        "й": "y",
        "к": "k",
        "л": "l",
        "м": "m",
        "н": "n",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "c",
        "ч": "ch",
        "ш": "sh",
        "щ": "sch",
        "ъ": "",
        "ы": "y",
        "ь": "",
        "э": "e",
        "ю": "yu",
        "я": "ya",
    }
    raw = "".join(translit.get(ch, ch) for ch in name.lower())
    raw = re.sub(r"[^a-z0-9]+", ".", raw).strip(".")
    return f"student.{raw}"[:80] or "student"


def grouped_student_names() -> dict[str, list[str]]:
    rows = defaultdict(set)
    schedules = CourseSchedule.query.join(User, CourseSchedule.teacher_id == User.id).all()
    for schedule in schedules:
        teacher_name = schedule.teacher.name
        for enrollment in schedule.enrollments:
            rows[teacher_name].add(enrollment.student.name)
    return {teacher: sorted(names) for teacher, names in sorted(rows.items())}

